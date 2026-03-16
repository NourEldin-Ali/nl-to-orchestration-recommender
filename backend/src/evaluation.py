from __future__ import annotations

import argparse
import json
import os
import re
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from time import perf_counter
from typing import Any, Iterable

PIPELINE_CONFIGS = {
    "one_step": {
        "one_step": True,
        "based_on_existing_orchestrator": False,
    },
    # "KB": {
    #     "one_step": False,
    #     "based_on_existing_orchestrator": False,
    # },
    "one_step_with_orchestrator": {
        "one_step": False,
        "based_on_existing_orchestrator": True,
    },
}

EXCEL_CELL_MAX_CHARS = 32000

DEFAULT_INPUTS = [
    "I need an open-source orchestrator to deploy my application in the cloud.",
    (
            "I am planning to prepare the infrastructure for my cloud application from "
            "scratch. I need a cloud orchestrator that supports multi-cloud deployment "
            "across AWS and Azure, and that can provision and configure the infrastructure."
        ),
        (
            "I am a doctoral student looking for recent and recognized cloud-edge "
            "orchestration frameworks."
        ),
         (
            "We already run Kubernetes in the cloud and want to extend orchestration to "
            "edge nodes, while supporting deployment, monitoring, and runtime control "
            "across cloud and edge."
        ),
        (
            "I am developing a telemedicine application for connected ambulances. The "
            "system collects real-time patient data from medical IoT sensors inside the "
            "ambulance. Critical data must be processed locally at the Edge for "
            "low-latency alerts, while selected information is sent to the Cloud for "
            "advanced analytics, storage, and remote monitoring by hospital doctors. I am "
            "therefore looking for an open source orchestration solution suitable to "
            "deploy and manage this multi-layer telemedicine application."
        ),
        (
            "I want one single tool that handles provisioning, configuration, service "
            "orchestration, workflow orchestration, and covers cloud, edge, and IoT."
        ),
]

RUN_HEADERS = [
    "run_id",
    "started_at_utc",
    "ended_at_utc",
    "llm_type",
    "model_name",
    "pipeline",
    "one_step",
    "based_on_existing_orchestrator",
    "repeat_index",
    "query_index",
    "thread_id",
    "user_query",
    "status",
    "final_response",
    "response_draft",
    "coverage",
    "recommendation_policy",
    "final_recommendation",
    "attempt_try",
    "human_turns",
    "input_tokens",
    "output_tokens",
    "total_tokens",
    "active_time_seconds",
    "wall_clock_seconds",
    "errors_json",
    "token_usage_by_node_json",
    "execution_timing_by_node_json",
    "state_json",
]


@dataclass(frozen=True)
class LLMRunConfig:
    llm_type: str
    model_name: str


DEFAULT_EVALUATION_LLM_CONFIGS = [
    LLMRunConfig(llm_type="nvidia", model_name="mistralai/mistral-large-3-675b-instruct-2512"),
    LLMRunConfig(llm_type="nvidia", model_name="mistralai/devstral-2-123b-instruct-2512"),
    LLMRunConfig(llm_type="nvidia", model_name="nvidia/nemotron-3-super-120b-a12b"),
    LLMRunConfig(llm_type="nvidia", model_name="meta/llama-3.3-70b-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="meta/llama-3.1-405b-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="qwen/qwen3.5-397b-a17b"),
    LLMRunConfig(llm_type="nvidia", model_name="qwen/qwen3-next-80b-a3b-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="qwen/qwen3.5-122b-a10b"),
    LLMRunConfig(llm_type="nvidia", model_name="deepseek-ai/deepseek-v3.2"),
    LLMRunConfig(llm_type="nvidia", model_name="openai/gpt-oss-120b"),
    LLMRunConfig(llm_type="nvidia", model_name="openai/gpt-oss-20b"),
    LLMRunConfig(llm_type="nvidia", model_name="meta/llama-3.2-3b-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="meta/llama-3.1-8b-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="mistralai/mistral-small-24b-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="mistralai/mistral-7b-instruct-v0.3"),
    LLMRunConfig(llm_type="nvidia", model_name="microsoft/phi-4-mini-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="microsoft/phi-3-small-128k-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="google/gemma-3-27b-it"),        
    LLMRunConfig(llm_type="nvidia", model_name="google/gemma-3-1b-it"),
    LLMRunConfig(llm_type="nvidia", model_name="qwen/qwen2.5-7b-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="qwen/qwen2-7b-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="microsoft/phi-3-small-128k-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="moonshotai/kimi-k2.5"),
    LLMRunConfig(llm_type="nvidia", model_name="moonshotai/kimi-k2-instruct"),
    LLMRunConfig(llm_type="nvidia", model_name="nvidia/nvidia-nemotron-nano-9b-v2"),
]


def parse_llm_spec(raw: str) -> LLMRunConfig:
    if ":" not in raw:
        raise argparse.ArgumentTypeError(
            f"Invalid --llm value '{raw}'. Expected format: llm_type:model_name"
        )
    llm_type, model_name = raw.split(":", maxsplit=1)
    llm_type = llm_type.strip()
    model_name = model_name.strip()
    if not llm_type or not model_name:
        raise argparse.ArgumentTypeError(
            f"Invalid --llm value '{raw}'. Both llm_type and model_name are required."
        )
    return LLMRunConfig(llm_type=llm_type, model_name=model_name)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Batch evaluation runner. Executes multiple prompts over multiple LLM "
            "configs and exports all run details to an Excel file."
        )
    )
    parser.add_argument(
        "--llm",
        dest="llms",
        action="append",
        default=[],
        help=(
            "LLM config in format llm_type:model_name. Repeat for multiple LLMs. "
            "If omitted, evaluation uses DEFAULT_EVALUATION_LLM_CONFIGS in src/evaluation.py "
            "(each entry contains both llm_type and model_name)."
        ),
    )
    parser.add_argument(
        "--input",
        dest="inputs",
        action="append",
        default=[],
        help="A user query to evaluate. Repeat this argument for multiple inputs.",
    )
    parser.add_argument(
        "--inputs-file",
        type=str,
        default="",
        help="Text file with one input query per line.",
    )
    parser.add_argument(
        "--pipelines",
        nargs="+",
        choices=sorted(PIPELINE_CONFIGS.keys()),
        default=["llm", "slm"],
        help=(
            "llm -> one_step baseline, "
            "slm -> full graph, "
            "llm_existing -> baseline grounded on detected existing orchestrators"
        ),
    )
    parser.add_argument(
        "--repeats",
        type=int,
        default=3,
        help="How many times to repeat each (LLM, input, pipeline) combination.",
    )
    parser.add_argument(
        "--auto-human-response",
        type=str,
        default="Yes, composition is acceptable if needed.",
        help="Automatic reply used when the full pipeline waits for human input.",
    )
    parser.add_argument(
        "--max-human-turns",
        type=int,
        default=2,
        help="Maximum automatic resumes when state is waiting_human.",
    )
    parser.add_argument(
        "--output",
        type=str,
        default="",
        help=(
            "Output .xlsx file name or path. The final file is always written inside "
            "--per-run-dir. Default file name: evaluation_results_YYYYMMDD_HHMMSS.xlsx"
        ),
    )
    parser.add_argument(
        "--per-run-dir",
        type=str,
        default="../output",
        help="Directory to save one detailed Excel file per run. Default: <output_stem>_per_run",
    )
    return parser.parse_args()


def resolve_llm_configs(raw_llm_specs: Iterable[str]) -> list[LLMRunConfig]:
    if raw_llm_specs:
        return [parse_llm_spec(spec) for spec in raw_llm_specs]

    if DEFAULT_EVALUATION_LLM_CONFIGS:
        return list(DEFAULT_EVALUATION_LLM_CONFIGS)

    # Fallback only if defaults are emptied manually.
    llm_type = (os.getenv("LLM_TYPE") or "groq").strip()
    model_name = (os.getenv("LLM_MODEL_NAME") or "llama-3.3-70b-versatile").strip()
    return [LLMRunConfig(llm_type=llm_type, model_name=model_name)]


def load_inputs(inline_inputs: Iterable[str], inputs_file: str) -> list[str]:
    queries: list[str] = []

    for query in inline_inputs:
        query = query.strip()
        if query:
            queries.append(query)

    if inputs_file:
        file_path = Path(inputs_file)
        if not file_path.exists():
            raise FileNotFoundError(f"inputs file not found: {file_path}")

        with file_path.open("r", encoding="utf-8") as f:
            for line in f:
                query = line.strip()
                if not query or query.startswith("#"):
                    continue
                queries.append(query)

    if not queries:
        return list(DEFAULT_INPUTS)
    return queries


@contextmanager
def temporary_llm_env(llm_type: str, model_name: str):
    previous = {
        "LLM_TYPE": os.getenv("LLM_TYPE"),
        "LLM_MODEL_NAME": os.getenv("LLM_MODEL_NAME"),
    }
    os.environ["LLM_TYPE"] = llm_type
    os.environ["LLM_MODEL_NAME"] = model_name
    try:
        yield
    finally:
        for key, old_value in previous.items():
            if old_value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = old_value


def run_graph_once(
    user_query: str,
    one_step: bool,
    based_on_existing_orchestrator: bool,
    thread_id: str,
    auto_human_response: str,
    max_human_turns: int,
) -> tuple[dict[str, Any], int]:
    from langchain_core.messages import HumanMessage
    from src.orcherstration_recommender.graph import build_graph

    graph = build_graph(
        one_step=one_step,
        based_on_existing_orchestrator=based_on_existing_orchestrator,
    )
    config = {"configurable": {"thread_id": thread_id}}
    state = graph.invoke(
        {
            "user_query": user_query,
            "messages": [],
            "one_step": one_step,
            "based_on_existing_orchestrator": based_on_existing_orchestrator,
        },
        config=config,
    )

    if not isinstance(state, dict):
        return (
            {
                "status": "failed",
                "errors": [f"Unexpected state type: {type(state).__name__}"],
                "final_response": "",
            },
            0,
        )

    human_turns = 0
    while state.get("status") == "waiting_human" and human_turns < max_human_turns:
        human_turns += 1
        graph.update_state(
            config,
            {"messages": [HumanMessage(content=auto_human_response)]},
        )
        state = graph.invoke(None, config=config)

    if state.get("status") == "waiting_human" and human_turns >= max_human_turns:
        errors = list(state.get("errors", []))
        errors.append(f"max_human_turns_reached={max_human_turns}")
        state = dict(state)
        state["errors"] = errors

    return state, human_turns


def json_string(value: Any) -> str:
    try:
        return json.dumps(value, ensure_ascii=False, default=str)
    except Exception:
        return json.dumps(str(value), ensure_ascii=False)


def to_float_or_none(value: Any) -> float | None:
    try:
        if value is None:
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def build_run_row(
    *,
    run_id: int,
    started_at: datetime,
    ended_at: datetime,
    llm_cfg: LLMRunConfig,
    pipeline: str,
    repeat_index: int,
    query_index: int,
    thread_id: str,
    user_query: str,
    state: dict[str, Any],
    human_turns: int,
    wall_clock_seconds: float,
) -> dict[str, Any]:
    pipeline_config = PIPELINE_CONFIGS[pipeline]
    token_usage = state.get("token_usage", {}) if isinstance(state, dict) else {}
    if not isinstance(token_usage, dict):
        token_usage = {}

    token_totals = token_usage.get("totals", {})
    if not isinstance(token_totals, dict):
        token_totals = {}

    execution_timing = state.get("execution_timing", {}) if isinstance(state, dict) else {}
    if not isinstance(execution_timing, dict):
        execution_timing = {}

    errors = state.get("errors", []) if isinstance(state, dict) else []
    if not isinstance(errors, list):
        errors = [str(errors)]

    return {
        "run_id": run_id,
        "started_at_utc": started_at.isoformat(),
        "ended_at_utc": ended_at.isoformat(),
        "llm_type": llm_cfg.llm_type,
        "model_name": llm_cfg.model_name,
        "pipeline": pipeline.upper(),
        "one_step": pipeline_config["one_step"],
        "based_on_existing_orchestrator": pipeline_config["based_on_existing_orchestrator"],
        "repeat_index": repeat_index,
        "query_index": query_index,
        "thread_id": thread_id,
        "user_query": user_query,
        "status": state.get("status", ""),
        "final_response": str(state.get("final_response", "")),
        "response_draft": str(state.get("response_draft", "")),
        "coverage": state.get("coverage", ""),
        "recommendation_policy": state.get("recommendation_policy", ""),
        "final_recommendation": state.get("final_recommendation", ""),
        "attempt_try": state.get("attempt_try", ""),
        "human_turns": human_turns,
        "input_tokens": token_totals.get("input_tokens", 0),
        "output_tokens": token_totals.get("output_tokens", 0),
        "total_tokens": token_totals.get("total_tokens", 0),
        "active_time_seconds": to_float_or_none(execution_timing.get("total_active_seconds")),
        "wall_clock_seconds": round(wall_clock_seconds, 6),
        "errors_json": json_string(errors),
        "token_usage_by_node_json": json_string(token_usage.get("by_node", {})),
        "execution_timing_by_node_json": json_string(execution_timing.get("by_node", {})),
        "state_json": json_string(state),
    }


def set_sheet_layout(ws):
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wide_columns = {
        "user_query",
        "final_response",
        "response_draft",
        "errors_json",
        "token_usage_by_node_json",
        "execution_timing_by_node_json",
        "state_json",
    }

    for col_idx, header in enumerate(RUN_HEADERS, start=1):
        width = 58 if header in wide_columns else 20
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wrap_cols = {RUN_HEADERS.index(name) + 1 for name in wide_columns}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_summary_sheet(workbook, run_rows: list[dict[str, Any]]) -> None:
    from openpyxl.utils import get_column_letter

    ws = workbook.create_sheet("summary")
    headers = [
        "llm_type",
        "model_name",
        "pipeline",
        "runs",
        "done_count",
        "failed_count",
        "waiting_human_count",
        "avg_wall_clock_seconds",
        "avg_total_tokens",
    ]
    ws.append(headers)

    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in run_rows:
        key = (row["llm_type"], row["model_name"], row["pipeline"])
        if key not in grouped:
            grouped[key] = {
                "runs": 0,
                "done_count": 0,
                "failed_count": 0,
                "waiting_human_count": 0,
                "wall_clock_sum": 0.0,
                "total_tokens_sum": 0.0,
            }

        g = grouped[key]
        g["runs"] += 1
        status = str(row.get("status", "")).lower()
        if status == "done":
            g["done_count"] += 1
        elif status == "failed":
            g["failed_count"] += 1
        elif status == "waiting_human":
            g["waiting_human_count"] += 1

        g["wall_clock_sum"] += float(row.get("wall_clock_seconds") or 0.0)
        g["total_tokens_sum"] += float(row.get("total_tokens") or 0.0)

    for (llm_type, model_name, pipeline), g in grouped.items():
        runs = g["runs"] or 1
        ws.append(
            [
                llm_type,
                model_name,
                pipeline,
                g["runs"],
                g["done_count"],
                g["failed_count"],
                g["waiting_human_count"],
                round(g["wall_clock_sum"] / runs, 6),
                round(g["total_tokens_sum"] / runs, 2),
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 26


def safe_filename_fragment(value: str, max_len: int = 40) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value)).strip("_")
    if not cleaned:
        cleaned = "value"
    return cleaned[:max_len]


def parse_json_field(raw: Any, fallback: Any) -> Any:
    if not isinstance(raw, str):
        return fallback
    try:
        return json.loads(raw)
    except Exception:
        return fallback


def write_run_detail_excel(output_dir: Path, row: dict[str, Any]) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to export Excel files. Install it with "
            "`pip install -r backend/requirements.txt`."
        ) from exc

    output_dir.mkdir(parents=True, exist_ok=True)

    file_name = (
        f"run_{int(row.get('run_id', 0)):05d}_"
        f"{safe_filename_fragment(row.get('pipeline', 'run'))}_"
        f"{safe_filename_fragment(row.get('llm_type', 'llm'))}_"
        f"{safe_filename_fragment(row.get('model_name', 'model'))}.xlsx"
    )
    run_path = output_dir / file_name

    workbook = Workbook()

    overview_ws = workbook.active
    overview_ws.title = "overview"
    overview_ws.append(["field", "value"])

    overview_fields = [
        "run_id",
        "started_at_utc",
        "ended_at_utc",
        "llm_type",
        "model_name",
        "pipeline",
        "one_step",
        "based_on_existing_orchestrator",
        "repeat_index",
        "query_index",
        "thread_id",
        "status",
        "coverage",
        "recommendation_policy",
        "final_recommendation",
        "attempt_try",
        "human_turns",
        "input_tokens",
        "output_tokens",
        "total_tokens",
        "active_time_seconds",
        "wall_clock_seconds",
        "user_query",
        "response_draft",
        "final_response",
    ]
    for field in overview_fields:
        overview_ws.append([field, row.get(field, "")])

    for header_cell in overview_ws[1]:
        header_cell.font = Font(bold=True)
    overview_ws.freeze_panes = "A2"
    overview_ws.column_dimensions["A"].width = 30
    overview_ws.column_dimensions["B"].width = 110
    for cells in overview_ws.iter_rows(min_row=2, max_row=overview_ws.max_row):
        cells[1].alignment = Alignment(wrap_text=True, vertical="top")

    errors_ws = workbook.create_sheet("errors")
    errors_ws.append(["index", "error"])
    errors = parse_json_field(row.get("errors_json"), fallback=[])
    if not isinstance(errors, list):
        errors = [str(errors)]
    if not errors:
        errors_ws.append([1, ""])
    else:
        for idx, err in enumerate(errors, start=1):
            errors_ws.append([idx, str(err)])
    for header_cell in errors_ws[1]:
        header_cell.font = Font(bold=True)
    errors_ws.freeze_panes = "A2"
    errors_ws.column_dimensions["A"].width = 12
    errors_ws.column_dimensions["B"].width = 120
    for cells in errors_ws.iter_rows(min_row=2, max_row=errors_ws.max_row):
        cells[1].alignment = Alignment(wrap_text=True, vertical="top")

    token_ws = workbook.create_sheet("token_by_node")
    token_ws.append(["node", "input_tokens", "output_tokens", "total_tokens"])
    token_by_node = parse_json_field(row.get("token_usage_by_node_json"), fallback={})
    if isinstance(token_by_node, dict) and token_by_node:
        for node_name, usage in token_by_node.items():
            usage_dict = usage if isinstance(usage, dict) else {}
            token_ws.append(
                [
                    str(node_name),
                    usage_dict.get("input_tokens", 0),
                    usage_dict.get("output_tokens", 0),
                    usage_dict.get("total_tokens", 0),
                ]
            )
    else:
        token_ws.append(["", 0, 0, 0])
    for header_cell in token_ws[1]:
        header_cell.font = Font(bold=True)
    token_ws.freeze_panes = "A2"
    token_ws.column_dimensions["A"].width = 45
    token_ws.column_dimensions["B"].width = 20
    token_ws.column_dimensions["C"].width = 20
    token_ws.column_dimensions["D"].width = 20

    timing_ws = workbook.create_sheet("timing_by_node")
    timing_ws.append(["node", "calls", "last_seconds", "total_seconds", "avg_seconds"])
    timing_by_node = parse_json_field(row.get("execution_timing_by_node_json"), fallback={})
    if isinstance(timing_by_node, dict) and timing_by_node:
        for node_name, usage in timing_by_node.items():
            usage_dict = usage if isinstance(usage, dict) else {}
            timing_ws.append(
                [
                    str(node_name),
                    usage_dict.get("calls", 0),
                    usage_dict.get("last_seconds", 0),
                    usage_dict.get("total_seconds", 0),
                    usage_dict.get("avg_seconds", 0),
                ]
            )
    else:
        timing_ws.append(["", 0, 0, 0, 0])
    for header_cell in timing_ws[1]:
        header_cell.font = Font(bold=True)
    timing_ws.freeze_panes = "A2"
    timing_ws.column_dimensions["A"].width = 45
    timing_ws.column_dimensions["B"].width = 15
    timing_ws.column_dimensions["C"].width = 18
    timing_ws.column_dimensions["D"].width = 18
    timing_ws.column_dimensions["E"].width = 18

    state_ws = workbook.create_sheet("state_json")
    state_ws.append(["chunk_index", "state_json_chunk"])
    state_json = str(row.get("state_json", ""))
    if not state_json:
        state_ws.append([1, ""])
    else:
        chunk_index = 1
        for start in range(0, len(state_json), EXCEL_CELL_MAX_CHARS):
            state_ws.append([chunk_index, state_json[start:start + EXCEL_CELL_MAX_CHARS]])
            chunk_index += 1
    for header_cell in state_ws[1]:
        header_cell.font = Font(bold=True)
    state_ws.freeze_panes = "A2"
    state_ws.column_dimensions["A"].width = 14
    state_ws.column_dimensions["B"].width = 120
    for cells in state_ws.iter_rows(min_row=2, max_row=state_ws.max_row):
        cells[1].alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(run_path)
    return run_path


def write_excel(output_path: Path, run_rows: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to export Excel files. Install it with "
            "`pip install -r backend/requirements.txt`."
        ) from exc

    workbook = Workbook()
    runs_ws = workbook.active
    runs_ws.title = "runs"
    runs_ws.append(RUN_HEADERS)

    for row in run_rows:
        runs_ws.append([row.get(header) for header in RUN_HEADERS])

    set_sheet_layout(runs_ws)
    build_summary_sheet(workbook, run_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_output_path(raw_output: str) -> Path:
    if raw_output.strip():
        return Path(raw_output).expanduser().resolve()
    now = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return Path.cwd() / f"evaluation_results_{now}.xlsx"


def build_per_run_dir(output_path: Path, raw_per_run_dir: str) -> Path:
    if raw_per_run_dir.strip():
        return Path(raw_per_run_dir).expanduser().resolve()
    return output_path.parent / f"{output_path.stem}_per_run"


def main() -> None:
    args = parse_args()

    if args.repeats < 1:
        raise ValueError("--repeats must be >= 1")
    if args.max_human_turns < 0:
        raise ValueError("--max-human-turns must be >= 0")

    llm_configs = resolve_llm_configs(args.llms)
    input_queries = load_inputs(args.inputs, args.inputs_file)
    pipelines = list(dict.fromkeys([pipeline.lower() for pipeline in args.pipelines]))
    requested_output_path = build_output_path(args.output)
    per_run_dir = build_per_run_dir(requested_output_path, args.per_run_dir)
    output_path = per_run_dir / requested_output_path.name

    total_runs = len(llm_configs) * len(input_queries) * len(pipelines) * args.repeats
    print(f"Starting evaluation with {total_runs} run(s)...")

    run_rows: list[dict[str, Any]] = []
    run_id = 0

    for llm_cfg in llm_configs:
        for pipeline in pipelines:
            pipeline_config = PIPELINE_CONFIGS[pipeline]
            one_step = pipeline_config["one_step"]
            based_on_existing_orchestrator = pipeline_config["based_on_existing_orchestrator"]
            for repeat_index in range(1, args.repeats + 1):
                for query_index, user_query in enumerate(input_queries, start=1):
                    run_id += 1
                    thread_id = (
                        f"eval-{run_id}-{pipeline}-{llm_cfg.llm_type}-{repeat_index}-{query_index}"
                    )

                    print(
                        f"[{run_id}/{total_runs}] "
                        f"llm={llm_cfg.llm_type}:{llm_cfg.model_name} "
                        f"pipeline={pipeline.upper()} "
                        f"query={query_index}/{len(input_queries)} repeat={repeat_index}"
                    )

                    started_at = datetime.now(timezone.utc)
                    wall_start = perf_counter()
                    human_turns = 0
                    state: dict[str, Any]

                    try:
                        with temporary_llm_env(llm_cfg.llm_type, llm_cfg.model_name):
                            state, human_turns = run_graph_once(
                                user_query=user_query,
                                one_step=one_step,
                                based_on_existing_orchestrator=based_on_existing_orchestrator,
                                thread_id=thread_id,
                                auto_human_response=args.auto_human_response,
                                max_human_turns=args.max_human_turns,
                            )
                    except Exception as exc:
                        state = {
                            "status": "failed",
                            "errors": [f"evaluation_runtime_error: {type(exc).__name__}: {exc}"],
                            "final_response": "",
                        }

                    wall_clock_seconds = perf_counter() - wall_start
                    ended_at = datetime.now(timezone.utc)

                    run_row = build_run_row(
                        run_id=run_id,
                        started_at=started_at,
                        ended_at=ended_at,
                        llm_cfg=llm_cfg,
                        pipeline=pipeline,
                        repeat_index=repeat_index,
                        query_index=query_index,
                        thread_id=thread_id,
                        user_query=user_query,
                        state=state,
                        human_turns=human_turns,
                        wall_clock_seconds=wall_clock_seconds,
                    )
                    run_rows.append(run_row)
                    write_run_detail_excel(per_run_dir, run_row)

    write_excel(output_path, run_rows)
    print(f"Evaluation finished. Global results saved to: {output_path}")
    print(f"Per-run detailed files saved to: {per_run_dir}")


if __name__ == "__main__":
    main()
