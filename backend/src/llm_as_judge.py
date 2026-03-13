from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class LLMRunConfig:
    llm_type: str
    model_name: str


DEFAULT_JUDGE_LLM_CONFIGS = [
    LLMRunConfig(llm_type="nvidia", model_name="qwen/qwen2-7b-instruct"),
    # LLMRunConfig(llm_type="openai", model_name="gpt-4.1-mini"),
]

# Edit this list directly.
DEFAULT_GROUND_TRUTH = [
    {
        "user_query": "I need an open-source orchestrator to deploy my application in the cloud.",
        "expected_answer": (
            "According to the knowledge base, Kubernetes is recommended as it satisfies "
            "the required orchestration capabilities, including deployment support and "
            "Cloud layer coverage. Would you like to know more about its adoption in "
            "practice?"
        ),
    },
    {
        "user_query": (
            "I am planning to prepare the infrastructure for my cloud application from "
            "scratch. I need a cloud orchestrator that supports multi-cloud deployment "
            "across AWS and Azure, and that can provision and configure the infrastructure."
        ),
        "expected_answer": (
            "Based on the knowledge base analysis, Cloudify is recommended. It supports "
            "infrastructure provisioning and configuration and enables multi-cloud "
            "deployment across AWS and Azure."
        ),
    },
    {
        "user_query": (
            "I am a doctoral student looking for recent and recognized cloud-edge "
            "orchestration frameworks."
        ),
        "expected_answer": (
            "According to the knowledge base, two relevant solutions match your intent, "
            "particularly with respect to recent and recognized cloud-edge orchestration "
            "frameworks. As a doctoral student, you may be especially interested in "
            "solutions with strong scientific contributions published in well-ranked "
            "conferences and journals. The two identified candidates are GAIKube and "
            "Oakestra. GAIKube is a recent, research-oriented framework, while Oakestra "
            "demonstrates stronger community adoption and practical maturity."
        ),
    },
    {
        "user_query": (
            "We already run Kubernetes in the cloud and want to extend orchestration to "
            "edge nodes, while supporting deployment, monitoring, and runtime control "
            "across cloud and edge."
        ),
        "expected_answer": (
            "Based on the analysis of your orchestration requirements, we identified two "
            "potential candidates capable of extending orchestration from Cloud to Edge: "
            "KubeEdge and Oakestra. Both orchestrators provide Edge support and satisfy "
            "the required capabilities for deployment, monitoring, and runtime control "
            "across distributed environments. However, based on your intent and your "
            "existing Kubernetes deployment, KubeEdge is the most suitable option, as it "
            "extends orchestration to the Edge while remaining fully aligned with the "
            "Kubernetes ecosystem."
        ),
    },
    {
        "user_query": (
            "I am developing a telemedicine application for connected ambulances. The "
            "system collects real-time patient data from medical IoT sensors inside the "
            "ambulance. Critical data must be processed locally at the Edge for "
            "low-latency alerts, while selected information is sent to the Cloud for "
            "advanced analytics, storage, and remote monitoring by hospital doctors. I am "
            "therefore looking for an open source orchestration solution suitable to "
            "deploy and manage this multi-layer telemedicine application."
        ),
        "expected_answer": (
            "Given your telemedicine scenario, we propose a composed orchestration "
            "architecture where each tool covers a specific part of the continuum. As you "
            "requested, we explain here which tool covers which part of the needed "
            "solution: Terraform is used first to provision the required infrastructure "
            "resources (e.g., cloud instances, networking, and any edge-host resources). "
            "Next, Ansible applies the necessary system and middleware configuration on "
            "top of the provisioned resources (OS configuration, packages, services, "
            "security hardening, and bootstrap steps), ensuring your environments are "
            "consistent and reproducible. For orchestration of your distributed "
            "application services across the cloud and edge layers, Kubernetes manages "
            "the cloud-side microservices (deployment, scaling, service management), "
            "while KubeEdge extends Kubernetes-native orchestration capabilities to the "
            "edge nodes inside the ambulance, enabling Edge deployment and coordination "
            "while staying compatible with the Kubernetes ecosystem. Finally, because this "
            "solution combines multiple tools, Apache Airflow is introduced as the "
            "coordination layer to sequence and automate the overall workflow - e.g., "
            "provision -> configure -> deploy in Cloud -> deploy at Edge -> "
            "validation/operations - so you can execute the full process reliably and "
            "repeatedly."
        ),
    },
    {
        "user_query": (
            "I want one single tool that handles provisioning, configuration, service "
            "orchestration, workflow orchestration, and covers cloud, edge, and IoT."
        ),
        "expected_answer": (
            "No single orchestrator in the knowledge base satisfies all these requirements."
        ),
    },
]

CRITERIA = [
    "technical_correctness",
    "requirement_coverage",
    "explanation_groundedness",
    "architectural_integration",
    "completeness",
]

RATING_LABEL = {
    1: "Weak",
    2: "Acceptable",
    3: "Excellent",
}

JUDGE_PROMPT = """
You are evaluating orchestration-tool recommendations against a gold standard.

Score each criterion with:
1 = Weak
2 = Acceptable
3 = Excellent

Criteria:
- technical_correctness: whether the recommendation is technically appropriate.
- requirement_coverage: whether the recommendation satisfies the user requirements.
- explanation_groundedness: whether the explanation justifies the recommendation using orchestration capabilities.
- architectural_integration: for composition scenarios, whether the proposed tools form a functional orchestration architecture. For single-tool scenarios, assess whether the solution is architecturally coherent.
- completeness: whether all tools expected by the gold standard are correctly identified.

Rules:
- Use the expected answer as the gold standard.
- Penalize hallucinated tools, unsupported claims, and recommendations outside the project tool set.
- Missing tools or capabilities implied by the expected answer must reduce completeness.
- If the answer is empty, failed, or unusable, all scores must be 1.
- Output JSON only.

Return exactly:
{
  "technical_correctness": {"score": 1, "justification": "..."},
  "requirement_coverage": {"score": 1, "justification": "..."},
  "explanation_groundedness": {"score": 1, "justification": "..."},
  "architectural_integration": {"score": 1, "justification": "..."},
  "completeness": {"score": 1, "justification": "..."},
  "overall_summary": "..."
}
""".strip()


def parse_llm_spec(raw: str) -> LLMRunConfig:
    if ":" not in raw:
        raise argparse.ArgumentTypeError(
            f"Invalid --llm value '{raw}'. Expected format: llm_type:model_name"
        )
    llm_type, model_name = raw.split(":", maxsplit=1)
    llm_type = llm_type.strip()
    model_name = model_name.strip()
    if not llm_type or not model_name:
        raise argparse.ArgumentTypeError(
            f"Invalid --llm value '{raw}'. Both llm_type and model_name are required."
        )
    return LLMRunConfig(llm_type=llm_type, model_name=model_name)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Score an evaluation workbook with one or more judge LLMs."
    )
    parser.add_argument(
        "--input-excel",
        type=str,
        default=r"../output/evaluation_results_20260312_092447.xlsx",
        help="Evaluation workbook path. Default: latest evaluation_results_*.xlsx in ../output",
    )
    parser.add_argument("--sheet-name", type=str, default="runs", help="Default: runs")
    parser.add_argument(
        "--llm",
        dest="llms",
        action="append",
        default=[],
        help="Judge model in format llm_type:model_name. Repeat for multiple models.",
    )
    parser.add_argument("--output", type=str, default="", help="Default: <input>_judged.xlsx")
    parser.add_argument("--max-rows", type=int, default=0, help="Optional debug limit.")
    return parser.parse_args()


def resolve_llm_configs(raw_llms: list[str]) -> list[LLMRunConfig]:
    if raw_llms:
        return [parse_llm_spec(raw) for raw in raw_llms]
    return list(DEFAULT_JUDGE_LLM_CONFIGS)


def normalize_query(text: Any) -> str:
    normalized = str(text or "").casefold()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def build_ground_truth_map() -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for item in DEFAULT_GROUND_TRUTH:
        query = normalize_query(item.get("user_query"))
        if not query:
            raise ValueError("Each DEFAULT_GROUND_TRUTH entry needs 'user_query'.")
        if not str(item.get("expected_answer") or "").strip():
            raise ValueError("Each DEFAULT_GROUND_TRUTH entry needs 'expected_answer'.")
        if query in result:
            raise ValueError(f"Duplicate ground truth query: {item['user_query']}")
        result[query] = item
    return result


def load_runs(workbook_path: Path, sheet_name: str) -> tuple[Any, list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {workbook_path}")

    sheet = workbook[sheet_name]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    rows: list[dict[str, Any]] = []

    for excel_row, values in enumerate(
        sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True),
        start=2,
    ):
        row = {header: value for header, value in zip(headers, values)}
        if not any(value not in (None, "") for value in row.values()):
            continue
        row["_excel_row"] = excel_row
        rows.append(row)

    return workbook, rows


def validate_ground_truth(run_rows: list[dict[str, Any]], ground_truth_map: dict[str, dict[str, Any]]) -> None:
    missing: list[str] = []
    seen: set[str] = set()
    for row in run_rows:
        query = normalize_query(row.get("user_query"))
        if not query or query in seen:
            continue
        seen.add(query)
        if query not in ground_truth_map:
            missing.append(str(row.get("user_query", "")).strip())

    if missing:
        preview = "\n".join(f"- {item}" for item in missing[:10])
        raise ValueError(
            "Missing entries in DEFAULT_GROUND_TRUTH for these workbook queries:\n" + preview
        )


def build_output_path(input_excel: Path, raw_output: str) -> Path:
    if raw_output.strip():
        return Path(raw_output).expanduser().resolve()
    return input_excel.with_name(f"{input_excel.stem}_judged.xlsx")


def resolve_input_excel(raw_input: str) -> Path:
    if raw_input.strip():
        input_excel = Path(raw_input).expanduser().resolve()
        if not input_excel.exists():
            raise FileNotFoundError(f"Input workbook not found: {input_excel}")
        return input_excel

    output_dir = (Path(__file__).resolve().parents[2] / "output").resolve()
    candidates = sorted(
        [
            path
            for path in output_dir.glob("evaluation_results_*.xlsx")
            if "_judged" not in path.stem
        ],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(
            "No default evaluation workbook found in ../output. "
            "Use --input-excel explicitly."
        )
    return candidates[0]


def auto_weak_result(reason: str) -> dict[str, Any]:
    return {
        criterion: {
            "score": 1,
            "justification": reason,
        }
        for criterion in CRITERIA
    } | {"overall_summary": reason}


def extract_json_object(text: str) -> str:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("Judge model did not return JSON.")
    return cleaned[start : end + 1]


def validate_judge_response(data: dict[str, Any]) -> dict[str, Any]:
    for criterion in CRITERIA:
        item = data.get(criterion)
        if not isinstance(item, dict):
            raise ValueError(f"Missing criterion '{criterion}' in judge output.")
        score = item.get("score")
        justification = str(item.get("justification", "")).strip()
        if score not in (1, 2, 3):
            raise ValueError(f"Invalid score for '{criterion}': {score}")
        if not justification:
            raise ValueError(f"Missing justification for '{criterion}'.")

    summary = str(data.get("overall_summary", "")).strip()
    if not summary:
        raise ValueError("Missing overall_summary in judge output.")

    return data


def cleanup_extracted_text(text: str) -> str:
    cleaned = text.strip().rstrip(",").strip()
    cleaned = re.sub(r'"\s*}\s*,?\s*$', "", cleaned, flags=re.S)
    cleaned = re.sub(r'"\s*$', "", cleaned, flags=re.S)
    cleaned = cleaned.replace('\\"', '"')
    cleaned = cleaned.replace("\\n", "\n")
    return cleaned.strip()


def salvage_judge_response(text: str) -> dict[str, Any]:
    payload = extract_json_object(text)
    ordered_keys = CRITERIA + ["overall_summary"]
    result: dict[str, Any] = {}

    for index, criterion in enumerate(CRITERIA):
        start = payload.find(f'"{criterion}"')
        next_key = ordered_keys[index + 1]
        end = payload.find(f'"{next_key}"', start + 1)
        if start == -1 or end == -1:
            raise ValueError(f"Could not salvage criterion '{criterion}'.")

        block = payload[start:end]
        score_match = re.search(r'"score"\s*:\s*([123])', block)
        justification_match = re.search(r'"justification"\s*:\s*"', block)
        if not score_match or not justification_match:
            raise ValueError(f"Could not salvage criterion '{criterion}'.")

        justification = cleanup_extracted_text(block[justification_match.end() :])
        result[criterion] = {
            "score": int(score_match.group(1)),
            "justification": justification,
        }

    summary_start = payload.find('"overall_summary"')
    summary_match = re.search(r'"overall_summary"\s*:\s*"', payload[summary_start:])
    if summary_start == -1 or not summary_match:
        raise ValueError("Could not salvage overall_summary.")

    summary_text = payload[summary_start + summary_match.end() :]
    result["overall_summary"] = cleanup_extracted_text(summary_text)
    return validate_judge_response(result)


def parse_judge_response(text: str) -> dict[str, Any]:
    payload = extract_json_object(text)
    try:
        return validate_judge_response(json.loads(payload))
    except json.JSONDecodeError:
        return salvage_judge_response(payload)


def response_to_text(response: Any) -> str:
    content = getattr(response, "content", response)
    if isinstance(content, str):
        return content.strip()
    if isinstance(content, list):
        parts: list[str] = []
        for item in content:
            if isinstance(item, str):
                parts.append(item)
            elif isinstance(item, dict) and "text" in item:
                parts.append(str(item["text"]))
            else:
                parts.append(str(item))
        return "\n".join(part.strip() for part in parts if part and str(part).strip())
    return str(content).strip()


def judge_with_retry(llm: Any, payload: dict[str, Any]) -> dict[str, Any]:
    from langchain_core.messages import AIMessage, HumanMessage, SystemMessage

    messages = [
        SystemMessage(content=JUDGE_PROMPT),
        HumanMessage(content=json.dumps(payload, ensure_ascii=False, indent=2)),
    ]

    response = llm.invoke(messages)
    raw_text = response_to_text(response)
    try:
        return parse_judge_response(raw_text)
    except Exception as first_error:
        retry = llm.invoke(
            messages
            + [
                AIMessage(content=raw_text),
                HumanMessage(content="Return valid JSON only. Keep the exact schema."),
            ]
        )
        retry_text = response_to_text(retry)
        try:
            return parse_judge_response(retry_text)
        except Exception as second_error:
            print(
                "  warning: judge output could not be parsed; "
                "falling back to Weak scores."
            )
            return auto_weak_result(
                f"Judge output parse failure. First error: {first_error}. "
                f"Second error: {second_error}."
            )


def overall_score(result: dict[str, Any]) -> float:
    return round(sum(result[key]["score"] for key in CRITERIA) / len(CRITERIA), 2)


def overall_rating(score: float) -> str:
    if score < 1.5:
        return RATING_LABEL[1]
    if score < 2.5:
        return RATING_LABEL[2]
    return RATING_LABEL[3]


def build_llm(llm_cfg: LLMRunConfig) -> Any:
    from src.config.llm_config import LLMConnector

    connector = LLMConnector(
        model_name=llm_cfg.model_name,
        llm_type=llm_cfg.llm_type,
        temperature=0.0,
    )
    return connector()


def judge_runs(
    run_rows: list[dict[str, Any]],
    judge_llms: list[LLMRunConfig],
    ground_truth_map: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []

    for judge_cfg in judge_llms:
        print(f"Judge model: {judge_cfg.llm_type}:{judge_cfg.model_name}")
        llm = build_llm(judge_cfg)
        cache: dict[str, dict[str, Any]] = {}

        for index, run_row in enumerate(run_rows, start=1):
            print(f"  [{index}/{len(run_rows)}] run_id={run_row.get('run_id')}")
            query = normalize_query(run_row.get("user_query"))
            ground_truth = ground_truth_map[query]

            answer = str(run_row.get("final_response") or "").strip()
            status = str(run_row.get("status") or "").strip().lower()

            if status != "done" or not answer:
                result = auto_weak_result(
                    f"Run status='{status or 'unknown'}' or final_response is empty."
                )
            else:
                payload = {
                    "scenario": {
                        "user_query": run_row.get("user_query"),
                    },
                    "gold_standard": {"expected_answer": ground_truth.get("expected_answer", "")},
                    "candidate": {
                        "pipeline": run_row.get("pipeline"),
                        "candidate_model": run_row.get("model_name"),
                        "response": answer,
                    },
                }

                cache_key = json.dumps(payload, ensure_ascii=False, sort_keys=True)
                if cache_key not in cache:
                    cache[cache_key] = judge_with_retry(llm, payload)
                result = cache[cache_key]

            score = overall_score(result)
            results.append(
                {
                    "run_id": run_row.get("run_id"),
                    "pipeline": run_row.get("pipeline"),
                    "candidate_model": run_row.get("model_name"),
                    "user_query": run_row.get("user_query"),
                    "judge_llm_type": judge_cfg.llm_type,
                    "judge_model_name": judge_cfg.model_name,
                    "technical_correctness_score": result["technical_correctness"]["score"],
                    "technical_correctness_justification": result["technical_correctness"][
                        "justification"
                    ],
                    "requirement_coverage_score": result["requirement_coverage"]["score"],
                    "requirement_coverage_justification": result["requirement_coverage"][
                        "justification"
                    ],
                    "explanation_groundedness_score": result["explanation_groundedness"][
                        "score"
                    ],
                    "explanation_groundedness_justification": result[
                        "explanation_groundedness"
                    ]["justification"],
                    "architectural_integration_score": result["architectural_integration"][
                        "score"
                    ],
                    "architectural_integration_justification": result[
                        "architectural_integration"
                    ]["justification"],
                    "completeness_score": result["completeness"]["score"],
                    "completeness_justification": result["completeness"]["justification"],
                    "overall_score": score,
                    "overall_rating": overall_rating(score),
                    "overall_summary": result["overall_summary"],
                }
            )

    return results


def remove_sheet(workbook: Any, name: str) -> None:
    if name in workbook.sheetnames:
        workbook.remove(workbook[name])


def write_judge_runs_sheet(workbook: Any, judge_rows: list[dict[str, Any]]) -> None:
    remove_sheet(workbook, "judge_runs")
    sheet = workbook.create_sheet("judge_runs")

    headers = [
        "run_id",
        "pipeline",
        "candidate_model",
        "user_query",
        "judge_llm_type",
        "judge_model_name",
        "technical_correctness_score",
        "technical_correctness_justification",
        "requirement_coverage_score",
        "requirement_coverage_justification",
        "explanation_groundedness_score",
        "explanation_groundedness_justification",
        "architectural_integration_score",
        "architectural_integration_justification",
        "completeness_score",
        "completeness_justification",
        "overall_score",
        "overall_rating",
        "overall_summary",
    ]
    sheet.append(headers)

    for row in judge_rows:
        sheet.append([row.get(header, "") for header in headers])

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    wrap_headers = {
        "user_query",
        "technical_correctness_justification",
        "requirement_coverage_justification",
        "explanation_groundedness_justification",
        "architectural_integration_justification",
        "completeness_justification",
        "overall_summary",
    }

    for index, header in enumerate(headers, start=1):
        width = 18
        if header in wrap_headers:
            width = 50
        sheet.column_dimensions[get_column_letter(index)].width = width

    wrap_indexes = {
        index for index, header in enumerate(headers, start=1) if header in wrap_headers
    }
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            if cell.column in wrap_indexes:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_judge_summary_sheet(workbook: Any, judge_rows: list[dict[str, Any]]) -> None:
    remove_sheet(workbook, "judge_summary")
    sheet = workbook.create_sheet("judge_summary")

    headers = [
        "judge_llm_type",
        "judge_model_name",
        "pipeline",
        "candidate_model",
        "runs",
        "avg_overall_score",
        "avg_technical_correctness",
        "avg_requirement_coverage",
        "avg_explanation_groundedness",
        "avg_architectural_integration",
        "avg_completeness",
    ]
    sheet.append(headers)

    grouped: dict[tuple[str, str, str, str], dict[str, float]] = {}
    for row in judge_rows:
        key = (
            str(row["judge_llm_type"]),
            str(row["judge_model_name"]),
            str(row["pipeline"]),
            str(row["candidate_model"]),
        )
        if key not in grouped:
            grouped[key] = {
                "runs": 0.0,
                "overall_score": 0.0,
                "technical_correctness": 0.0,
                "requirement_coverage": 0.0,
                "explanation_groundedness": 0.0,
                "architectural_integration": 0.0,
                "completeness": 0.0,
            }

        group = grouped[key]
        group["runs"] += 1
        group["overall_score"] += float(row["overall_score"])
        group["technical_correctness"] += float(row["technical_correctness_score"])
        group["requirement_coverage"] += float(row["requirement_coverage_score"])
        group["explanation_groundedness"] += float(row["explanation_groundedness_score"])
        group["architectural_integration"] += float(row["architectural_integration_score"])
        group["completeness"] += float(row["completeness_score"])

    for key, group in grouped.items():
        runs = max(group["runs"], 1)
        sheet.append(
            [
                key[0],
                key[1],
                key[2],
                key[3],
                int(group["runs"]),
                round(group["overall_score"] / runs, 2),
                round(group["technical_correctness"] / runs, 2),
                round(group["requirement_coverage"] / runs, 2),
                round(group["explanation_groundedness"] / runs, 2),
                round(group["architectural_integration"] / runs, 2),
                round(group["completeness"] / runs, 2),
            ]
        )

    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 24


def write_judge_config_sheet(workbook: Any, input_excel: Path, judge_llms: list[LLMRunConfig]) -> None:
    remove_sheet(workbook, "judge_config")
    sheet = workbook.create_sheet("judge_config")
    sheet.append(["field", "value"])
    sheet.append(["input_excel", str(input_excel)])
    sheet.append(["generated_at", datetime.utcnow().isoformat()])
    sheet.append(["judge_models", json.dumps([cfg.__dict__ for cfg in judge_llms])])
    sheet.append(["ground_truth_entries", len(DEFAULT_GROUND_TRUTH)])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 100


def main() -> None:
    args = parse_args()
    if args.max_rows < 0:
        raise ValueError("--max-rows must be >= 0")

    input_excel = resolve_input_excel(args.input_excel)

    judge_llms = resolve_llm_configs(args.llms)
    ground_truth_map = build_ground_truth_map()
    workbook, run_rows = load_runs(input_excel, args.sheet_name)

    if args.max_rows:
        run_rows = run_rows[: args.max_rows]

    validate_ground_truth(run_rows, ground_truth_map)

    judge_rows = judge_runs(run_rows, judge_llms, ground_truth_map)
    write_judge_runs_sheet(workbook, judge_rows)
    write_judge_summary_sheet(workbook, judge_rows)
    write_judge_config_sheet(workbook, input_excel, judge_llms)

    output_path = build_output_path(input_excel, args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Judged workbook saved to: {output_path}")


if __name__ == "__main__":
    main()
